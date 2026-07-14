"""
02_ratio_analysis.py
=====================
Loads Waaree Energies' consolidated financial statements (FY20-FY26,
sourced from Part I / Screener.in), computes a full institutional ratio
suite, and produces:
  - output/ratio_trends.png       (4-panel chart)
  - output/ratios.csv
  - output/Waaree_Ratios.xlsx     (formatted workbook)

No internet required — runs entirely on the embedded historical dataset.

Usage:
    python 02_ratio_analysis.py
"""
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

HERE = os.path.dirname(__file__)
OUT_DIR = os.path.join(HERE, "output")
os.makedirs(OUT_DIR, exist_ok=True)

YEARS = ["FY20","FY21","FY22","FY23","FY24","FY25","FY26"]

# ---- Embedded historicals (₹ Crore, Consolidated) — Source: Part I / Screener.in ----
data = {
    "Revenue":            [1996,1953,2854,6751,11398,14444,26537],
    "EBITDA":             [93,85,106,836,1574,2722,5909],
    "Depreciation":       [27,29,43,164,277,402,990],
    "Interest":           [34,31,41,82,140,152,280],
    "OtherIncome":        [25,45,97,88,576,398,413],
    "NetProfit":          [39,48,80,500,1274,1928,3884],
    "EquityCapital":      [197,197,197,243,263,287,288],
    "Reserves":           [101,148,231,1595,3825,9192,14150],
    "Borrowings":         [157,340,363,320,553,1199,3213],
    "OtherLiabilities":   [483,596,1362,5247,6635,9028,12465],
    "NetFixedAssets":     [152,285,625,1105,1450,4051,7329],
    "CWIP":               [40,3,124,537,1341,1884,3476],
    "Investments":        [85,115,143,31,71,65,953],
    "OtherAssets":        [661,878,1262,5732,8414,13706,18358],
    "CFO":                [83,67,698,1560,2305,3158,1627],
    "FCF":                [32,-35,202,698,968,-114,-3209],
}
df = pd.DataFrame(data, index=YEARS)
df["EBIT"] = df["EBITDA"] - df["Depreciation"]
df["PBT"] = df["EBIT"] + df["OtherIncome"] - df["Interest"]
df["TotalEquity"] = df["EquityCapital"] + df["Reserves"]
df["TotalAssets"] = df["NetFixedAssets"] + df["CWIP"] + df["Investments"] + df["OtherAssets"]
df["CapitalEmployed"] = df["TotalEquity"] + df["Borrowings"]

# ---- Ratio computations ----
ratios = pd.DataFrame(index=YEARS)
ratios["Revenue Growth %"] = df["Revenue"].pct_change() * 100
ratios["EBITDA Margin %"] = df["EBITDA"] / df["Revenue"] * 100
ratios["Net Margin %"] = df["NetProfit"] / df["Revenue"] * 100
ratios["ROE %"] = df["NetProfit"] / df["TotalEquity"].rolling(2).mean() * 100
ratios["ROCE %"] = df["EBIT"] / df["CapitalEmployed"].rolling(2).mean() * 100
ratios["Debt/Equity (x)"] = df["Borrowings"] / df["TotalEquity"]
ratios["Interest Coverage (x)"] = df["EBIT"] / df["Interest"]
ratios["Asset Turnover (x)"] = df["Revenue"] / df["TotalAssets"]
ratios["CFO/EBITDA (x)"] = df["CFO"] / df["EBITDA"]
ratios["FCF Margin %"] = df["FCF"] / df["Revenue"] * 100

ratios.to_csv(os.path.join(OUT_DIR, "ratios.csv"))
print("Computed ratio suite:\n")
print(ratios.round(2).to_string())

# ---- 4-panel chart ----
fig, axes = plt.subplots(2, 2, figsize=(12, 8))
fig.suptitle("Waaree Energies — Financial Trend Analysis (FY20-FY26)", fontsize=14, fontweight="bold", color="#1F2A44")

ax = axes[0, 0]
ax.bar(YEARS, df["Revenue"], color="#1F2A44", label="Revenue (₹ Cr)")
ax2 = ax.twinx()
ax2.plot(YEARS, ratios["EBITDA Margin %"], color="#B08D57", marker="o", linewidth=2, label="EBITDA Margin %")
ax.set_title("Revenue Growth & EBITDA Margin")
ax.set_ylabel("₹ Crore")
ax2.set_ylabel("EBITDA Margin %")
ax2.yaxis.set_major_formatter(mticker.PercentFormatter())

ax = axes[0, 1]
ax.plot(YEARS, ratios["ROE %"], marker="o", color="#1F2A44", linewidth=2, label="ROE %")
ax.plot(YEARS, ratios["ROCE %"], marker="s", color="#B08D57", linewidth=2, label="ROCE %")
ax.set_title("Return on Equity vs. Return on Capital Employed")
ax.set_ylabel("%")
ax.legend()
ax.grid(alpha=0.3)

ax = axes[1, 0]
ax.bar(YEARS, ratios["Debt/Equity (x)"], color="#7A6C5D")
ax.set_title("Leverage — Debt / Equity (x)")
ax.set_ylabel("x")

ax = axes[1, 1]
ax.bar(YEARS, df["FCF"], color=["#A32424" if v < 0 else "#1E6B3C" for v in df["FCF"]])
ax.axhline(0, color="black", linewidth=0.8)
ax.set_title("Free Cash Flow (₹ Cr) — the FY25-FY26 cash-conversion flag")
ax.set_ylabel("₹ Crore")

plt.tight_layout(rect=[0, 0, 1, 0.96])
chart_path = os.path.join(OUT_DIR, "ratio_trends.png")
plt.savefig(chart_path, dpi=150)
plt.close()
print(f"\nSaved chart -> {chart_path}")

# ---- Excel export ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

wb = Workbook()
ws = wb.active
ws.title = "Ratios"
ws["A1"] = "Waaree Energies — Ratio Analysis (Python-computed, FY20-FY26)"
ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F2A44")
ws.merge_cells("A1:H1")

ws.append([])
header = ["Metric"] + YEARS
ws.append(header)
for c in ws[3]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F2A44")
for metric in ratios.columns:
    row = [metric] + [None if pd.isna(v) else round(v, 2) for v in ratios[metric]]
    ws.append(row)
for col in "ABCDEFGHI":
    ws.column_dimensions[col].width = 20

ws2 = wb.create_sheet("Chart")
img = XLImage(chart_path)
img.width, img.height = 720, 480
ws2.add_image(img, "A1")

xlsx_path = os.path.join(OUT_DIR, "Waaree_Ratios.xlsx")
wb.save(xlsx_path)
print(f"Saved workbook -> {xlsx_path}")
